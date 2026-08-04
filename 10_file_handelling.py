# File Handelling:

    # Using open()  we can create, read, write files.


# file1 = open("./my_files/file1.txt", "w")
# file2 = open("./my_files/file2.csv", "w")
# file3 = open("./my_files/file3.py", "w")
# file4 = open("./my_files/file4.js", "w")
# file5 = open("./my_files/file5.java", "w")
# file6 = open("./my_files/file6.xlsx", "w")
# file7 = open("./my_files/file7.c", "w")


# ------------ Create a file and Write Content in it ----------

# file1 = open("./my_files/file1.txt", "w")
# file1.write("Hello Python developers..!")
# file1.close()


# file1 = open("./my_files/file1.txt", "w")
# file1.write("This is my first file.")
# file1.close()


# file1 = open("./my_files/file1.txt", "a")
# file1.write("\nHello Python Developers..!")
# file1.close()

# file1 = open("./my_files/file1.txt", "r")
# content = file1.read()

# print(content)
# file1.close()

#------------------------------------

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "employees"

ws["A1"] = "Firstname"
ws["B1"] = "Lastname"
ws["C1"] = "Age"
ws["D1"] = "Salary"
ws["E1"] = "Email"

data = [
    ["Ajay", "Yadav", 34, 50000, 'ajay.yadav@gmail.com'],
    ["Ajay", "Yadav", 34, 50000, 'ajay.yadav@gmail.com'],
    ["Ajay", "Yadav", 34, 50000, 'ajay.yadav@gmail.com'],
    ["Ajay", "Yadav", 34, 50000, 'ajay.yadav@gmail.com'],
    ["Ajay", "Yadav", 34, 50000, 'ajay.yadav@gmail.com'],
    ["Ajay", "Yadav", 34, 50000, 'ajay.yadav@gmail.com']

]

for row in data:
    ws.append(row)

wb.save("./my_files/employees.csv")